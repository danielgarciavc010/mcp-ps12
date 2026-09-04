"""
MCP Server - ivanti-itsm
Transporte: stdio

Arranque:
    python server.py
    -> servidor MCP por stdin/stdout

Herramientas:
    - list_business_objects
    - get_business_object
    - create_business_object
    - create_incident_simple
    - update_business_object
    - delete_business_object
    - export_to_excel
    - run_quick_action
    - list_fields
    - run_saved_search
    - get_related_objects
    - link_business_objects
    - get_metadata
    - find_user_info
    - list_service_catalog_templates
    - get_template_fields
    - get_field_valid_values
    - create_service_request
    - upload_attachment

Referencia de la API:
https://help.ivanti.com/ht/help/en_US/ISM/2021/admin/Content/Configure/API/RestAPI-Introduction.htm
"""

import os
import re
import sys
import unicodedata
import pathlib as _pathlib
from datetime import datetime
from enum import Enum
from typing import Any, Optional

import httpx
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

# Config
load_dotenv()

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stdin, "reconfigure"):
    sys.stdin.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

TENANT_URL = os.getenv("IVANTI_TENANT_URL", "").rstrip("/")
API_KEY    = os.getenv("IVANTI_API_KEY", "")
TIMEOUT    = 30.0

if not TENANT_URL or not API_KEY:
    raise RuntimeError(
        "Configura IVANTI_TENANT_URL e IVANTI_API_KEY en el archivo .env "
        "(copia .env.example como .env y rellena la API Key)."
    )

HEADERS = {
    "Authorization": f"rest_api_key={API_KEY}",
    "Content-Type":  "application/json",
}


class IncidentLevel(str, Enum):
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"

mcp = FastMCP(
    name="ivanti-itsm",
    instructions=(
        "Herramientas para gestionar Ivanti Neurons ITSM: consultar, crear, "
        "actualizar y eliminar business objects (incidentes, solicitudes de "
        "servicio, problemas, cambios, empleados), ejecutar quick actions, "
        "gestionar solicitudes del catalogo de servicios y subir adjuntos."
    ),
)


def _clean_unicode(obj: Any) -> Any:
    if isinstance(obj, str):
        s = obj.replace(chr(160), " ").replace("\u00a0", " ")
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        return s.encode("ascii", errors="replace").decode("ascii")
    elif isinstance(obj, dict):
        return {_clean_unicode(k): _clean_unicode(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_clean_unicode(elem) for elem in obj]
    elif isinstance(obj, tuple):
        return tuple(_clean_unicode(elem) for elem in obj)
    return obj


def _error(code: str, message: str) -> dict:
    return {"ok": False, "error": {"code": code, "message": message}}


def _odata_literal(value: Any) -> str:
    if value is None:
        return "null"

    if isinstance(value, bool):
        return str(value).lower()

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(value)

    str_value = str(value).strip()

    # Detectar fecha y hora ISO 8601.
    # Ejemplos aceptados:
    # 2026-08-02T00:00:00Z
    # 2026-08-02T00:00:00.000Z
    # 2026-08-02T02:00:00+02:00
    if re.fullmatch(
        r"\d{4}-\d{2}-\d{2}"
        r"T\d{2}:\d{2}:\d{2}"
        r"(?:\.\d{1,7})?"
        r"(?:Z|[+-]\d{2}:\d{2})",
        str_value,
    ):
        return str_value

    escaped = str_value.replace("'", "''")
    return f"'{escaped}'"


def _build_odata_filter(filters: list[dict[str, Any]]) -> str | None:
    allowed_operators = {
        "eq", "ne", "gt", "ge", "lt", "le",
        "contains", "startswith", "endswith",
    }
    expressions = []
    for item in filters:
        field = str(item.get("field", "")).strip()
        operator = str(item.get("operator", "")).lower().strip()
        if (
            not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", field)
            or operator not in allowed_operators
            or "value" not in item
        ):
            return None
        value = _odata_literal(item["value"])
        if operator in {"contains", "startswith", "endswith"}:
            expressions.append(f"{operator}({field}, {value})")
        else:
            expressions.append(f"{field} {operator} {value}")
    return " and ".join(f"({expression})" for expression in expressions)


def _build_odata_order_by(order_by: str) -> str | None:
    match = re.fullmatch(
        r"([A-Za-z_][A-Za-z0-9_]*)(?:\s+(asc|desc))?",
        order_by.strip(),
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    field, direction = match.groups()
    return f"{field} {(direction or 'asc').lower()}"


async def _request(
    method: str,
    endpoint: str,
    *,
    headers: dict | None = None,
    **kwargs,
) -> tuple[Any, dict | None]:
    req_headers = headers if headers is not None else HEADERS
    try:
        async with httpx.AsyncClient(
            base_url=TENANT_URL,
            headers=req_headers,
            timeout=TIMEOUT,
        ) as client:
            response = await client.request(method, endpoint, **kwargs)
    except httpx.TimeoutException:
        return None, _error("TIMEOUT", "El servicio no respondio a tiempo.")
    except httpx.RequestError as exc:
        return None, _error(
            "CONNECTION_ERROR", f"No se pudo conectar al servicio: {exc}."
        )

    if response.status_code >= 400:
        return None, _error(
            "HTTP_ERROR",
            f"Ivanti API error {response.status_code}: {response.text}",
        )

    try:
        return _clean_unicode(response.json()), None
    except Exception:
        return _clean_unicode(response.text), None


# Tools - Business Objects (CRUD)

@mcp.tool()
async def list_business_objects(
    object_name: str,
    top: int = 10,
    select: Optional[str] = None,
    search: Optional[str] = None,
    filters: Optional[list[dict[str, Any]]] = None,
    order_by: Optional[str] = None,
) -> dict:
    """
    Lista registros de un business object de Ivanti (p.ej. 'incidents',
    'employees', 'serviceReqs', 'problems', 'changes'). El nombre del
    objeto siempre va en plural.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').
        top: numero maximo de registros a devolver (max. 25).
        select: campos a devolver separados por coma (ej: "Subject,Status").
        search: palabra clave de busqueda de texto libre ($search).
        filters: filtros estructurados. Cada elemento debe incluir
            "field", "operator" y "value". Operadores: eq, ne, gt, ge,
            lt, le, contains, startswith, endswith. Se combinan con AND.
        order_by: campo y direccion de ordenacion OData, por ejemplo
            "CreatedDateTime desc". Si no se indica direccion, usa asc.

    Returns:
        {"ok": true, "data": { ...registros del servicio... }}
    """
    DEFAULT_SELECT = (
        "RecId,IncidentNumber,Subject,Status,Priority,"
        "Category,CreatedBy,CreatedDateTime"
    )
    EMPLOYEE_DEFAULT_SELECT = (
        "RecId,DisplayName,LoginID,PrimaryEmail,Department,Status,"
        "Title,ManagerEmail,EmployeeLocation,BusinessUnit,PrimaryPhone,Disabled"
    )
    MAX_TOP = 25

    if filters:
        generated_filter = _build_odata_filter(filters)
        if generated_filter is None:
            return _error(
                "INVALID_FILTER",
                "Cada filtro requiere field, operator valido y value.",
            )
    else:
        generated_filter = None

    generated_order_by = None
    if order_by:
        generated_order_by = _build_odata_order_by(order_by)
        if generated_order_by is None:
            return _error(
                "INVALID_ORDER_BY",
                "order_by debe tener el formato 'Campo asc|desc'.",
            )

    params: dict[str, Any] = {
        "$top": min(top, MAX_TOP),
        "$count": "true",
    }
    if generated_filter:
        params["$filter"] = generated_filter
    if generated_order_by:
        params["$orderby"] = generated_order_by
    if select:
        params["$select"] = select
    elif object_name.lower() in ("incidents", "incident"):
        params["$select"] = DEFAULT_SELECT
    elif object_name.lower() in ("employees", "employee"):
        params["$select"] = EMPLOYEE_DEFAULT_SELECT
    if search:
        params["$search"] = search

    raw, err = await _request(
        "GET", f"/api/odata/businessobject/{object_name}", params=params
    )
    if err:
        return err

    return {"ok": True, "data": raw}


@mcp.tool()
async def get_business_object(object_name: str, rec_id: str) -> dict:
    """
    Obtiene un registro concreto de un business object por su RecId.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').
        rec_id: RecId (identificador unico) del registro.

    Returns:
        {"ok": true, "data": { ...campos del registro... }}
    """
    raw, err = await _request(
        "GET", f"/api/odata/businessobject/{object_name}('{rec_id}')"
    )
    if err:
        return err

    return {"ok": True, "data": raw}


@mcp.tool()
async def create_business_object(
    object_name: str, fields: dict[str, Any]
) -> dict:
    """
    Crea un nuevo registro de un business object.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').
        fields: diccionario con los campos y valores a establecer, ej:
            {"Subject": "No imprime", "Status": "Open", "Priority": 3}

    Returns:
        {"ok": true, "data": { ...registro creado... }}
    """
    raw, err = await _request(
        "POST", f"/api/odata/businessobject/{object_name}", json=fields
    )
    if err:
        return err

    return {"ok": True, "data": raw}


@mcp.tool()
async def create_incident_simple(
    subject: str,
    description: str,
    customer_email: str,
    urgency: IncidentLevel = IncidentLevel.MEDIUM,
    impact: IncidentLevel = IncidentLevel.MEDIUM,
) -> dict:
    """
    Crea un ticket de incidencia de forma segura, resolviendo las
    validaciones tecnicas de Ivanti. USALO siempre que el usuario pida
    crear un incidente o reportar un problema.

    Args:
        subject: Resumen breve del problema.
        description: Detalles completos del sintoma o problema.
        customer_email: Email del usuario afectado.
        urgency: Urgencia: 'High', 'Medium', 'Low'.
        impact: Impacto: 'High', 'Medium', 'Low'.

    Returns:
        {"ok": true, "data": {"IncidentNumber": "...", "RecId": "...", ...}}
    """
    subject = subject.strip()
    description = description.strip()
    customer_email = customer_email.strip()
    if not subject or not description or not customer_email:
        return _error(
            "INVALID_INCIDENT_DATA",
            "subject, description y customer_email son obligatorios y no pueden estar vacios.",
        )

    if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", customer_email):
        return _error(
            "INVALID_CUSTOMER_EMAIL",
            "customer_email debe ser un email valido del usuario afectado.",
        )

    try:
        urgency = IncidentLevel(urgency)
        impact = IncidentLevel(impact)
    except (TypeError, ValueError):
        return _error(
            "INVALID_INCIDENT_DATA",
            "urgency e impact deben ser 'High', 'Medium' o 'Low'.",
        )

    safe_customer_email = customer_email.replace("'", "''")
    search_filter = f"PrimaryEmail eq '{safe_customer_email}'"

    raw_emp, err = await _request(
        "GET",
        "/api/odata/businessobject/employees",
        params={
            "$filter": search_filter,
            "$top": 1,
            "$select": "RecId,DisplayName,PrimaryEmail",
        },
    )
    if err:
        return _error(
            "EMPLOYEE_SEARCH_ERROR",
            f"Error buscando el email '{customer_email}': {err['error']['message']}",
        )

    values = raw_emp.get("value", []) if isinstance(raw_emp, dict) else []
    if not values:
        return _error(
            "EMPLOYEE_NOT_FOUND",
            f"No se encontro un empleado con el email '{customer_email}'. "
            "Pide al usuario que verifique el email.",
        )
    employee = values[0]

    customer_rec_id = employee.get("RecId")
    if not customer_rec_id:
        return _error(
            "EMPLOYEE_INVALID",
            f"El empleado '{customer_email}' no contiene un RecId valido.",
        )
    customer_display = employee.get("DisplayName", customer_email)

    fields = {
        "Subject": subject,
        "Symptom": description,
        "Urgency": urgency.value,
        "Impact": impact.value,
        "Status": "Logged",
        "ProfileLink_RecID": customer_rec_id,
        "ProfileLink_Category": "Employee",
    }

    raw, err = await _request(
        "POST", "/api/odata/businessobject/incidents", json=fields
    )
    if err:
        return err

    if not isinstance(raw, dict):
        return _error(
            "INCIDENT_CREATE_INVALID_RESPONSE",
            "Ivanti no devolvio los datos del incidente creado.",
        )

    inc_number = raw.get("IncidentNumber", "desconocido")
    return {
        "ok": True,
        "data": {
            "message": f"Incidente #{inc_number} creado correctamente para {customer_display}.",
            "IncidentNumber": inc_number,
            "RecId": raw.get("RecId", "") if isinstance(raw, dict) else "",
            "Status": raw.get("Status", "Logged") if isinstance(raw, dict) else "Logged",
        },
    }


@mcp.tool()
async def update_business_object(
    object_name: str, rec_id: str, fields: dict[str, Any]
) -> dict:
    """
    Actualiza campos de un registro existente.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').
        rec_id: RecId del registro a actualizar.
        fields: campos a modificar, ej: {"Status": "Closed"}

    Returns:
        {"ok": true, "data": { ...registro actualizado... }}
    """
    raw, err = await _request(
        "PATCH",
        f"/api/odata/businessobject/{object_name}('{rec_id}')",
        json=fields,
    )
    if err:
        return err

    return {"ok": True, "data": raw if raw else {"status": "ok"}}


@mcp.tool()
async def delete_business_object(object_name: str, rec_id: str) -> dict:
    """
    Elimina un registro de un business object.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').
        rec_id: RecId del registro a eliminar.

    Returns:
        {"ok": true, "data": {"status": "deleted", "rec_id": "..."}}
    """
    _, err = await _request(
        "DELETE", f"/api/odata/businessobject/{object_name}('{rec_id}')"
    )
    if err:
        return err

    return {"ok": True, "data": {"status": "deleted", "rec_id": rec_id}}

# Tools - Export

@mcp.tool()
async def export_to_excel(
    object_name: str,
    select: Optional[str] = None,
    search: Optional[str] = None,
    filters: Optional[list[dict[str, Any]]] = None,
    order_by: Optional[str] = None,
) -> dict:
    """
    Exporta TODOS los registros de un business object a un archivo Excel,
    con paginación automática. Genera un archivo .xlsx y devuelve el contenido
    dentro del workspace de la sesión.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').
        select: campos a exportar separados por coma (ej: "Subject,Status").
                Si no se especifica, usa los campos por defecto de list_business_objects.
        search: palabra clave de busqueda de texto libre ($search).
        filters: filtros estructurados. Cada elemento debe incluir
                "field", "operator" y "value". Operadores: eq, ne, gt, ge,
                lt, le, contains, startswith, endswith. Se combinan con AND.
        order_by: campo y direccion de ordenacion OData, por ejemplo
                "CreatedDateTime desc". Si no se indica direccion, usa asc.

    Returns:
        {
            "ok": true,
            "data": {
                "filename": "incidents_2026-09-02.xlsx",
                "record_count": 150
            }
        }
    """
    DEFAULT_SELECT = (
        "RecId,IncidentNumber,Subject,Status,Priority,"
        "Category,CreatedBy,CreatedDateTime"
    )
    EMPLOYEE_DEFAULT_SELECT = (
        "RecId,DisplayName,LoginID,PrimaryEmail,Department,Status,"
        "Title,ManagerEmail,EmployeeLocation,BusinessUnit,PrimaryPhone,Disabled"
    )
    PAGE_SIZE = 100  # Límite máximo de Ivanti API (ISM_4000)

    # Construir filtro OData
    if filters:
        generated_filter = _build_odata_filter(filters)
        if generated_filter is None:
            return _error(
                "INVALID_FILTER",
                "Cada filtro requiere field, operator valido y value.",
            )
    else:
        generated_filter = None

    # Construir order_by OData
    generated_order_by = None
    if order_by:
        generated_order_by = _build_odata_order_by(order_by)
        if generated_order_by is None:
            return _error(
                "INVALID_ORDER_BY",
                "order_by debe tener el formato 'Campo asc|desc'.",
            )

    # Determinar campos a usar
    if select:
        selected_fields = select
    elif object_name.lower() in ("incidents", "incident"):
        selected_fields = DEFAULT_SELECT
    elif object_name.lower() in ("employees", "employee"):
        selected_fields = EMPLOYEE_DEFAULT_SELECT
    else:
        selected_fields = None

    # Traer todos los registros con paginación
    all_records = []
    skip = 0

    while True:
        params: dict[str, Any] = {
            "$top": PAGE_SIZE,
            "$skip": skip,
        }
        if generated_filter:
            params["$filter"] = generated_filter
        if generated_order_by:
            params["$orderby"] = generated_order_by
        if selected_fields:
            params["$select"] = selected_fields
        if search:
            params["$search"] = search

        raw, err = await _request(
            "GET", f"/api/odata/businessobject/{object_name}", params=params
        )
        if err:
            return err

        # Extraer registros
        if isinstance(raw, dict) and "value" in raw:
            records = raw["value"]
        else:
            records = []

        if not records:
            break

        all_records.extend(records)
        skip += PAGE_SIZE

    if not all_records:
        return _error(
            "NO_DATA",
            "La consulta no devolvio ningun registro.",
        )

    # Crear Excel con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = object_name[:31]  # Los nombres de hojas en Excel tienen máx 31 caracteres

    # Encabezados
    if all_records:
        headers = list(all_records[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Datos
        for row_idx, record in enumerate(all_records, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = record.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Ajustar ancho de columnas
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for record in all_records:
                max_len = max(max_len, len(str(record.get(header, ""))))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    # Generar nombre del archivo
    today = datetime.utcnow().strftime("%Y-%m-%d")
    safe_object_name = re.sub(r"[^A-Za-z0-9_-]", "_", object_name).strip("._") or "export"
    filename = f"{safe_object_name}_{today}.xlsx"

    # El workflow ejecuta el MCP con el workspace de la sesión como cwd.
    file_path = os.path.join(os.getcwd(), filename)
    wb.save(file_path)

    return {
        "ok": True,
        "data": {
            "filename": filename,
            "record_count": len(all_records),
        },
    }

    
# Tools - Metadata & Discovery

# Valores conocidos para campos comunes (no siempre expuestos en metadatos XML)
KNOWN_FIELD_VALUES = {
    "Status": [
        {"name": "Logged", "description": ""},
        {"name": "Assigned", "description": ""},
        {"name": "Active", "description": ""},
        {"name": "Waiting for Resolution", "description": ""},
        {"name": "Resolved", "description": ""},
        {"name": "Closed", "description": ""},
    ],
    "Priority": [
        {"name": "1", "description": "Priority 1 - Critical"},
        {"name": "2", "description": "Priority 2 - High"},
        {"name": "3", "description": "Priority 3 - Medium"},
        {"name": "4", "description": "Priority 4 - Low"},
        {"name": "5", "description": "Priority 5 - Planning"},
    ],
    "Urgency": [
        {"name": "Low", "description": ""},
        {"name": "Medium", "description": ""},
        {"name": "High", "description": ""},
    ],
    "Impact": [
        {"name": "Low", "description": ""},
        {"name": "Medium", "description": ""},
        {"name": "High", "description": ""},
    ],
}

@mcp.tool()
async def get_metadata(object_name: str) -> dict:
    """
    Devuelve los metadatos (campos, tipos, relaciones...) de un business
    object como texto XML plano.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').

    Returns:
        {"ok": true, "data": "<xml>...metadatos...</xml>"}
    """
    async with httpx.AsyncClient(
        base_url=TENANT_URL, headers=HEADERS, timeout=TIMEOUT
    ) as client:
        try:
            resp = await client.get(f"/api/odata/{object_name}/$metadata")
        except httpx.TimeoutException:
            return _error("TIMEOUT", "El servicio no respondio a tiempo.")
        except httpx.RequestError as exc:
            return _error("CONNECTION_ERROR", f"No se pudo conectar: {exc}.")

    if resp.status_code >= 400:
        return _error("HTTP_ERROR", f"Ivanti API error {resp.status_code}: {resp.text}")

    raw_bytes = resp.content
    if raw_bytes[:2] in (b"\xff\xfe", b"\xfe\xff"):
        text = raw_bytes.decode("utf-16")
    else:
        text = resp.text

    return _clean_unicode({"ok": True, "data": text})


@mcp.tool()
async def list_fields(object_name: str) -> dict:
    """
    Devuelve la lista de campos (nombre y tipo) de un business object,
    parseada desde el XML de metadatos. Mas compacta que get_metadata().
    Para campos de tipo enumerado, incluye los valores permitidos.

    Args:
        object_name: nombre del business object en plural (ej: 'incidents').

    Returns:
        {"ok": true, "data": {"object_name": "...", "fields": [...]}}
    """
    result = await get_metadata(object_name)
    if not result.get("ok"):
        return result

    xml_text = result["data"]

    candidates = [object_name, object_name.rstrip("s")]
    entity_block = None
    matched_name = None

    for candidate in candidates:
        pattern = rf'<EntityType Name="{re.escape(candidate)}">'
        match = re.search(pattern, xml_text, flags=re.IGNORECASE)
        if match:
            start = match.start()
            end_pos = xml_text.find("</EntityType>", start)
            if end_pos == -1:
                continue
            end = end_pos + len("</EntityType>")
            entity_block = xml_text[start:end]
            matched_name = candidate
            break

    if entity_block is None:
        available = sorted(
            set(re.findall(r'<EntityType Name="([^"]+)"', xml_text))
        )
        return _error(
            "ENTITY_NOT_FOUND",
            f"No se encontro EntityType para '{object_name}'. "
            f"Tipos disponibles: {available[:50]}",
        )

    props = re.findall(
        r'<Property Name="([^"]+)" Type="([^"]+)"', entity_block
    )

    # Extraer todos los EnumTypes disponibles en los metadatos
    enum_types = {}
    try:
        # Buscar cada EnumType en el XML
        enum_start_pattern = r'<EnumType\s+Name="([^"]+)"'
        for enum_start_match in re.finditer(enum_start_pattern, xml_text):
            enum_name = enum_start_match.group(1)
            start_pos = enum_start_match.start()
           
            # Buscar el cierre </EnumType> correspondiente
            end_pattern = r'</EnumType>'
            end_match = re.search(end_pattern, xml_text[start_pos:])
            if not end_match:
                continue
           
            end_pos = start_pos + end_match.end()
            enum_content = xml_text[start_pos:end_pos]
           
            # Extraer los valores (Members) del enum
            members = []
            for member_match in re.finditer(
                r'<Member\s+Name="([^"]+)"(?:\s+Value="([^"]*)")?',
                enum_content
            ):
                member_name = member_match.group(1)
                member_value = member_match.group(2) or ""
                members.append({"name": member_name, "value": member_value})
           
            if members:  # Solo agregar si hay miembros
                enum_types[enum_name] = members
    except Exception:
        # Si falla la extraccion de enums, continuar sin ellos
        pass

    fields_list = []
    for name, type_str in props:
        field_data = {
            "name": name,
            "type": type_str,
        }

        # Si el tipo es un EnumType, incluir los valores permitidos
        # El tipo puede ser "Namespace.EnumTypeName" o solo "EnumTypeName"
        type_base = type_str.split(".")[-1] if "." in type_str else type_str
        if type_base in enum_types:
            field_data["enum_values"] = enum_types[type_base]
        # Tambien verificar valores conocidos hardcodeados (para campos que no vienen en XML)
        elif name in KNOWN_FIELD_VALUES:
            field_data["enum_values"] = KNOWN_FIELD_VALUES[name]

        fields_list.append(field_data)

    return {
        "ok": True,
        "data": {
            "object_name": object_name,
            "entity_type_matched": matched_name,
            "field_count": len(props),
            "fields": fields_list,
        },
    }


# Arranque
def main():
    mcp.run(transport="stdio")


if __name__ == "__main__":
    main()
